"""보고서 생성 (PDF + Excel)."""
import calendar
import io
from datetime import datetime, timedelta
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from models import Job, JobStatus, Printer, User


# ============================================================
# 한글 폰트 등록
# ============================================================

_FONT_REGISTERED = False


def _register_korean_font():
    global _FONT_REGISTERED
    if _FONT_REGISTERED:
        return
    pdfmetrics.registerFont(
        TTFont("NotoSans", "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc")
    )
    pdfmetrics.registerFont(
        TTFont("NotoSans-Bold", "/usr/share/fonts/opentype/noto/NotoSansCJK-Bold.ttc")
    )
    _FONT_REGISTERED = True


# ============================================================
# 데이터 수집
# ============================================================

async def gather_report_data(db: AsyncSession, year: int, month: int):
    """해당 월의 보고서 데이터 수집."""

    # 기간
    start = datetime(year, month, 1)
    if month == 12:
        end = datetime(year + 1, 1, 1)
    else:
        end = datetime(year, month + 1, 1)

    # 해당 기간의 모든 작업
    result = await db.execute(
        select(Job)
        .where(Job.created_at >= start)
        .where(Job.created_at < end)
        .order_by(Job.created_at.desc())
    )
    jobs = result.scalars().all()

    # 관련 사용자/프린터 정보
    user_ids = {j.user_id for j in jobs}
    printer_ids = {j.printer_id for j in jobs}

    users = {}
    if user_ids:
        result = await db.execute(select(User).where(User.id.in_(user_ids)))
        users = {u.id: u for u in result.scalars().all()}

    printers = {}
    if printer_ids:
        result = await db.execute(select(Printer).where(Printer.id.in_(printer_ids)))
        printers = {p.id: p for p in result.scalars().all()}

    # 통계
    total = len(jobs)
    completed = sum(1 for j in jobs if j.status == JobStatus.COMPLETED)
    failed = sum(1 for j in jobs if j.status == JobStatus.FAILED)
    rejected = sum(1 for j in jobs if j.status == JobStatus.REJECTED)
    pending = sum(
        1 for j in jobs
        if j.status in [JobStatus.PENDING_APPROVAL, JobStatus.QUEUED, JobStatus.PRINTING]
    )

    success_rate = (completed / total * 100) if total > 0 else 0
    unique_users = len(user_ids)

    # 프린터별 통계
    printer_stats = {}
    for printer_id in printer_ids:
        printer_jobs = [j for j in jobs if j.printer_id == printer_id]
        printer_stats[printer_id] = {
            "name": printers[printer_id].name if printer_id in printers else "—",
            "total": len(printer_jobs),
            "completed": sum(1 for j in printer_jobs if j.status == JobStatus.COMPLETED),
            "failed": sum(1 for j in printer_jobs if j.status == JobStatus.FAILED),
        }

    # 사용자별 출력 수 (Top 5)
    user_counts = {}
    for j in jobs:
        if j.status == JobStatus.COMPLETED:
            user_counts[j.user_id] = user_counts.get(j.user_id, 0) + 1
    top_users = sorted(user_counts.items(), key=lambda x: x[1], reverse=True)[:5]
    top_users_data = [
        {
            "name": users[uid].name if uid in users else "—",
            "email": users[uid].email if uid in users else "—",
            "count": count,
        }
        for uid, count in top_users
    ]

    return {
        "year": year,
        "month": month,
        "start": start,
        "end": end,
        "jobs": jobs,
        "users": users,
        "printers": printers,
        "stats": {
            "total": total,
            "completed": completed,
            "failed": failed,
            "rejected": rejected,
            "pending": pending,
            "success_rate": round(success_rate, 1),
            "unique_users": unique_users,
        },
        "printer_stats": printer_stats,
        "top_users": top_users_data,
    }


# ============================================================
# PDF 생성
# ============================================================

def _status_korean(status: str) -> str:
    return {
        "pending_approval": "승인 대기",
        "queued": "큐 대기",
        "printing": "출력 중",
        "completed": "완료",
        "failed": "실패",
        "rejected": "거부",
        "canceled": "취소",
    }.get(status, status)


def generate_pdf(data: dict) -> bytes:
    """월간 보고서 PDF 생성."""
    _register_korean_font()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
    )

    # 스타일
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title",
        parent=styles["Title"],
        fontName="NotoSans-Bold",
        fontSize=18,
        spaceAfter=8,
        alignment=1,  # center
    )
    h1_style = ParagraphStyle(
        "H1",
        parent=styles["Heading1"],
        fontName="NotoSans-Bold",
        fontSize=13,
        spaceBefore=18,
        spaceAfter=8,
        textColor=colors.HexColor("#1a1a1a"),
    )
    body_style = ParagraphStyle(
        "Body",
        parent=styles["BodyText"],
        fontName="NotoSans",
        fontSize=10,
        leading=14,
    )
    muted_style = ParagraphStyle(
        "Muted",
        parent=body_style,
        fontSize=9,
        textColor=colors.HexColor("#6b7280"),
        alignment=1,  # center
    )

    story = []

    # 헤더
    story.append(Paragraph(
        f"HAFS PrintQueue 운영 보고서",
        title_style,
    ))
    story.append(Paragraph(
        f"{data['year']}년 {data['month']}월",
        ParagraphStyle("Subtitle", parent=body_style, fontSize=14, alignment=1, spaceAfter=4),
    ))
    story.append(Paragraph(
        f"용인한국외국어대학교부설고등학교 메이커 시스템",
        muted_style,
    ))
    story.append(Spacer(1, 0.5 * cm))

    # 요약 통계
    story.append(Paragraph("요약 통계", h1_style))
    stats = data["stats"]
    summary_data = [
        ["항목", "값"],
        ["총 출력 신청", f"{stats['total']}건"],
        ["완료", f"{stats['completed']}건"],
        ["실패", f"{stats['failed']}건"],
        ["거부", f"{stats['rejected']}건"],
        ["진행 중", f"{stats['pending']}건"],
        ["성공률", f"{stats['success_rate']}%"],
        ["이용 학생 수", f"{stats['unique_users']}명"],
    ]
    summary_table = Table(summary_data, colWidths=[6 * cm, 6 * cm])
    summary_table.setStyle(TableStyle([
        ("FONT", (0, 0), (-1, -1), "NotoSans", 10),
        ("FONT", (0, 0), (-1, 0), "NotoSans-Bold", 10),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#374151")),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(summary_table)

    # 프린터별 통계
    if data["printer_stats"]:
        story.append(Paragraph("프린터별 통계", h1_style))
        printer_data = [["프린터", "신청 수", "완료", "실패"]]
        for pid, ps in data["printer_stats"].items():
            printer_data.append([
                ps["name"],
                str(ps["total"]),
                str(ps["completed"]),
                str(ps["failed"]),
            ])
        printer_table = Table(printer_data, colWidths=[5 * cm, 3 * cm, 3 * cm, 3 * cm])
        printer_table.setStyle(TableStyle([
            ("FONT", (0, 0), (-1, -1), "NotoSans", 10),
            ("FONT", (0, 0), (-1, 0), "NotoSans-Bold", 10),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
            ("ALIGN", (1, 1), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(printer_table)

    # 활발한 사용자 Top 5
    if data["top_users"]:
        story.append(Paragraph("가장 활발한 학생 (완료 기준 Top 5)", h1_style))
        top_data = [["순위", "이름", "이메일", "완료 수"]]
        for i, u in enumerate(data["top_users"], start=1):
            top_data.append([str(i), u["name"], u["email"], f"{u['count']}건"])
        top_table = Table(top_data, colWidths=[1.5 * cm, 4 * cm, 7 * cm, 2.5 * cm])
        top_table.setStyle(TableStyle([
            ("FONT", (0, 0), (-1, -1), "NotoSans", 10),
            ("FONT", (0, 0), (-1, 0), "NotoSans-Bold", 10),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
            ("ALIGN", (0, 1), (0, -1), "CENTER"),
            ("ALIGN", (3, 1), (3, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(top_table)

    # 푸터
    story.append(Spacer(1, 1 * cm))
    footer_text = f"생성 시각: {datetime.now().strftime('%Y-%m-%d %H:%M')} · HAFS PrintQueue · hafs.wiki"
    story.append(Paragraph(footer_text, muted_style))

    doc.build(story)
    pdf_bytes = buf.getvalue()
    buf.close()
    return pdf_bytes


# ============================================================
# Excel 생성
# ============================================================

def generate_excel(data: dict) -> bytes:
    """월간 작업 데이터 Excel 생성."""
    wb = Workbook()

    # === Sheet 1: 작업 목록 ===
    ws = wb.active
    ws.title = "작업 목록"

    headers = [
        "신청 일시", "학생 이름", "이메일", "프린터",
        "파일명", "상태", "승인 시각", "시작 시각", "완료 시각",
        "메모", "관리자 메모",
    ]
    ws.append(headers)

    # 헤더 스타일
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F46E5")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 데이터
    for job in data["jobs"]:
        user = data["users"].get(job.user_id)
        printer = data["printers"].get(job.printer_id)
        ws.append([
            job.created_at.strftime("%Y-%m-%d %H:%M") if job.created_at else "",
            user.name if user else "—",
            user.email if user else "—",
            printer.name if printer else "—",
            job.filename,
            _status_korean(job.status.value),
            job.approved_at.strftime("%Y-%m-%d %H:%M") if job.approved_at else "",
            job.started_at.strftime("%Y-%m-%d %H:%M") if job.started_at else "",
            job.completed_at.strftime("%Y-%m-%d %H:%M") if job.completed_at else "",
            job.user_notes or "",
            job.admin_notes or "",
        ])

    # 열 너비
    widths = [18, 12, 25, 12, 30, 10, 18, 18, 18, 25, 25]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    # === Sheet 2: 요약 ===
    ws2 = wb.create_sheet("요약")
    ws2.append([f"{data['year']}년 {data['month']}월 운영 보고"])
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.append([])
    ws2.append(["항목", "값"])
    for cell in ws2[3]:
        cell.font = header_font
        cell.fill = header_fill

    stats = data["stats"]
    rows = [
        ["총 출력 신청", f"{stats['total']}건"],
        ["완료", f"{stats['completed']}건"],
        ["실패", f"{stats['failed']}건"],
        ["거부", f"{stats['rejected']}건"],
        ["진행 중", f"{stats['pending']}건"],
        ["성공률", f"{stats['success_rate']}%"],
        ["이용 학생 수", f"{stats['unique_users']}명"],
    ]
    for r in rows:
        ws2.append(r)

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 20

    # 저장
    buf = io.BytesIO()
    wb.save(buf)
    excel_bytes = buf.getvalue()
    buf.close()
    return excel_bytes
